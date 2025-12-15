from io import BytesIO, StringIO
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from app.models.attendance import AttendanceRecord
from app.models.student import Student
from app.models.session import Session as SessionModel
from typing import List
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


class ReportService:
    """Service layer for report generation."""

    @staticmethod
    def generate_attendance_csv(
        db: Session, class_name: str = None, days: int = 30
    ) -> BytesIO:
        """Generate CSV attendance report."""
        cutoff_date = datetime.now() - timedelta(days=days)

        query = (
            db.query(
                Student.student_code,
                Student.first_name,
                Student.last_name,
                Student.class_name,
                AttendanceRecord.status,
                AttendanceRecord.marked_at,
                AttendanceRecord.late_minutes,
                AttendanceRecord.justification,
                AttendanceRecord.percentage,
            )
            .select_from(Student)
            .join(AttendanceRecord, AttendanceRecord.student_id == Student.id)
            .filter(AttendanceRecord.marked_at >= cutoff_date)
        )

        if class_name:
            query = query.filter(Student.class_name == class_name)

        records = query.order_by(Student.student_code, AttendanceRecord.marked_at).all()

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "Student Code",
                "First Name",
                "Last Name",
                "Class",
                "Status",
                "Date/Time",
                "Late Minutes",
                "Justification",
                "Percentage",
            ]
        )

        for record in records:
            writer.writerow(
                [
                    record.student_code,
                    record.first_name,
                    record.last_name,
                    record.class_name,
                    record.status,
                    record.marked_at.strftime("%Y-%m-%d %H:%M") if record.marked_at else "",
                    record.late_minutes or 0,
                    record.justification or "",
                    record.percentage or 0,
                ]
            )

        bytes_output = BytesIO()
        bytes_output.write(output.getvalue().encode("utf-8"))
        bytes_output.seek(0)
        return bytes_output

    @staticmethod
    def generate_attendance_summary(
        db: Session, student_id: int = None, class_name: str = None, days: int = 30
    ) -> dict:
        """Generate attendance summary statistics."""
        cutoff_date = datetime.now() - timedelta(days=days)

        query = (
            db.query(
                Student.id,
                Student.student_code,
                Student.first_name,
                Student.last_name,
                Student.class_name,
                Student.attendance_rate,
            )
            .distinct()
        )

        if student_id:
            query = query.filter(Student.id == student_id)
        elif class_name:
            query = query.filter(Student.class_name == class_name)

        students = query.all()

        summary = []
        for student in students:
            records = (
                db.query(AttendanceRecord)
                .filter(
                    AttendanceRecord.student_id == student.id,
                    AttendanceRecord.marked_at >= cutoff_date,
                )
                .all()
            )

            total = len(records)
            present = sum(1 for r in records if r.status == "present")
            absent = sum(1 for r in records if r.status == "absent")
            late = sum(1 for r in records if r.status == "late")
            excused = sum(1 for r in records if r.status == "excused")

            rate = (present / total * 100) if total > 0 else 0

            summary.append(
                {
                    "student_code": student.student_code,
                    "name": f"{student.first_name} {student.last_name}",
                    "class": student[4],  # class column from query
                    "total_sessions": total,
                    "present": present,
                    "absent": absent,
                    "late": late,
                    "excused": excused,
                    "attendance_rate": round(rate, 2),
                }
            )

        return {"report_date": datetime.now().isoformat(), "students": summary}

    @staticmethod
    def generate_student_report(db: Session, student_id: int) -> dict:
        """Generate comprehensive report for a student."""
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return None

        # Get attendance summary
        records = (
            db.query(AttendanceRecord)
            .filter(AttendanceRecord.student_id == student_id)
            .all()
        )

        total = len(records)
        present = sum(1 for r in records if r.status == "present")
        absent = sum(1 for r in records if r.status == "absent")
        late = sum(1 for r in records if r.status == "late")

        rate = (present / total * 100) if total > 0 else 0

        return {
            "student_code": student.student_code,
            "name": f"{student.first_name} {student.last_name}",
            "class": student.class_name,
            "email": student.email,
            "enrollment_date": student.enrollment_date.isoformat() if student.enrollment_date else None,
            "total_absence_hours": student.total_absence_hours,
            "total_late_minutes": student.total_late_minutes,
            "attendance_rate": student.attendance_rate,
            "alert_level": student.alert_level,
            "attendance_summary": {
                "total_sessions": total,
                "present": present,
                "absent": absent,
                "late": late,
                "rate": round(rate, 2),
            },
            "report_generated": datetime.now().isoformat(),
        }

    @staticmethod
    def generate_class_analytics(db: Session, class_name: str) -> dict:
        """Generate analytics for an entire class."""
        students = db.query(Student).filter(Student.class_name == class_name).all()

        if not students:
            return None

        total_students = len(students)
        high_risk_count = sum(1 for s in students if s.alert_level in ["high", "critical"])
        avg_attendance = sum(s.attendance_rate for s in students) / total_students

        return {
            "class": class_name,
            "total_students": total_students,
            "high_risk_students": high_risk_count,
            "average_attendance_rate": round(avg_attendance, 2),
            "students": [
                {
                    "code": s.student_code,
                    "name": f"{s.first_name} {s.last_name}",
                    "attendance_rate": s.attendance_rate,
                    "alert_level": s.alert_level,
                    "absence_hours": s.total_absence_hours,
                }
                for s in students
            ],
            "report_generated": datetime.now().isoformat(),
        }

    @staticmethod
    def generate_excel_attendance_report(db: Session, class_name: str = None) -> BytesIO:
        """Generate Excel attendance report with formatting."""
        cutoff_date = datetime.now() - timedelta(days=30)

        query = (
            db.query(
                Student.student_code,
                Student.first_name,
                Student.last_name,
                Student.class_name,
                AttendanceRecord.status,
                AttendanceRecord.marked_at,
                AttendanceRecord.late_minutes,
                AttendanceRecord.justification,
                AttendanceRecord.percentage,
            )
            .select_from(Student)
            .join(AttendanceRecord, AttendanceRecord.student_id == Student.id)
            .filter(AttendanceRecord.marked_at >= cutoff_date)
        )

        if class_name:
            query = query.filter(Student.class_name == class_name)

        records = query.order_by(Student.student_code, AttendanceRecord.marked_at).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Student Code",
            "First Name",
            "Last Name",
            "Class",
            "Status",
            "Date/Time",
            "Late (mins)",
            "Justification",
            "Percentage",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data rows
        status_fill_map = {
            "present": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "absent": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "late": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "excused": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        }

        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=record.student_code).border = border
            ws.cell(row=row_idx, column=2, value=record.first_name).border = border
            ws.cell(row=row_idx, column=3, value=record.last_name).border = border
            ws.cell(row=row_idx, column=4, value=record.class_name).border = border
            
            status_cell = ws.cell(row=row_idx, column=5, value=record.status)
            status_cell.border = border
            status_cell.fill = status_fill_map.get(record.status, PatternFill())
            
            ws.cell(
                row=row_idx,
                column=6,
                value=record.marked_at.strftime("%Y-%m-%d %H:%M") if record.marked_at else "",
            ).border = border
            ws.cell(row=row_idx, column=7, value=record.late_minutes or 0).border = border
            ws.cell(row=row_idx, column=8, value=record.justification or "").border = border
            ws.cell(row=row_idx, column=9, value=record.percentage or 0).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf_report(db: Session, student_id: int = None, class_name: str = None) -> BytesIO:
        """Generate PDF report using reportlab."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        # Title
        story.append(Paragraph("Smart Presence AI - Attendance Report", title_style))
        story.append(Paragraph(f"<b>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</b>", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        if student_id:
            report = ReportService.generate_student_report(db, student_id)
            if report:
                story.append(Paragraph(f"<b>Student:</b> {report['name']}", styles['Heading2']))
                story.append(Paragraph(f"<b>Code:</b> {report['student_code']}", styles['Normal']))
                story.append(Paragraph(f"<b>Class:</b> {report.get('class', 'N/A')}", styles['Normal']))
                story.append(Paragraph(f"<b>Email:</b> {report.get('email', 'N/A')}", styles['Normal']))
                story.append(Spacer(1, 0.2*inch))

                # Attendance Summary Table
                summary = report['attendance_summary']
                data = [
                    ['Metric', 'Value'],
                    ['Total Sessions', str(summary['total_sessions'])],
                    ['Present', str(summary['present'])],
                    ['Absent', str(summary['absent'])],
                    ['Late', str(summary['late'])],
                    ['Attendance Rate', f"{summary['rate']}%"],
                ]

                table = Table(data, colWidths=[2.5*inch, 2.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(table)
        else:
            report = ReportService.generate_class_analytics(db, class_name)
            if report:
                story.append(Paragraph(f"<b>Class:</b> {report['class']}", styles['Heading2']))
                story.append(Spacer(1, 0.2*inch))

                # Class Summary
                data = [
                    ['Metric', 'Value'],
                    ['Total Students', str(report['total_students'])],
                    ['High Risk Students', str(report['high_risk_students'])],
                    ['Average Attendance', f"{report['average_attendance_rate']}%"],
                ]

                table = Table(data, colWidths=[2.5*inch, 2.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(table)

        doc.build(story)
        output.seek(0)
        return output

