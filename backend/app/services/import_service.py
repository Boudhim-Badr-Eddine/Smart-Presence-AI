from __future__ import annotations

import csv
import io
from datetime import date, datetime, time
from typing import Dict, Iterable, List, Tuple

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.session import Session as SessionModel
from app.models.student import Student
from app.models.trainer import Trainer
from app.models.user import User
from app.services.auth import get_password_hash

# Lightweight importer to cover CSV/XLSX ingestion for admin bulk flows.


class ImportService:
    SUPPORTED_ENTITIES = {"students", "trainers", "sessions"}

    @staticmethod
    def _read_rows(upload: UploadFile) -> List[Dict[str, str]]:
        """Read CSV or XLSX into list of dicts with normalized headers."""
        name = (upload.filename or "").lower()
        data = upload.file.read()
        upload.file.seek(0)

        if name.endswith(".csv"):
            text = data.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return [ImportService._normalize_row(row) for row in reader]

        if name.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(data), read_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            normalized_headers = [ImportService._normalize_header(h) for h in headers]
            results: List[Dict[str, str]] = []
            for r in rows[1:]:
                row_dict = {
                    normalized_headers[i]: ("" if val is None else str(val).strip())
                    for i, val in enumerate(r)
                }
                results.append(ImportService._normalize_row(row_dict))
            return results

        raise ValueError("Unsupported file type. Use CSV or XLSX.")

    @staticmethod
    def _normalize_header(header: str) -> str:
        return header.lower().strip().replace(" ", "_")

    @staticmethod
    def _normalize_row(row: Dict[str, str]) -> Dict[str, str]:
        return {ImportService._normalize_header(k): (v or "").strip() for k, v in row.items()}

    @staticmethod
    def parse_upload(upload: UploadFile) -> List[Dict[str, str]]:
        return ImportService._read_rows(upload)

    # ---------- Entity specific importers ----------
    @staticmethod
    def _ensure_user(db: Session, email: str, first_name: str, last_name: str, role: str) -> User:
        user = db.query(User).filter(User.email == email).first()
        if user:
            return user
        user = User(
            username=email.split("@")[0],
            email=email,
            password_hash=get_password_hash("password123"),
            role=role,
            is_active=True,
        )
        db.add(user)
        db.flush()
        return user

    @staticmethod
    def import_students(db: Session, rows: Iterable[Dict[str, str]]) -> Tuple[int, List[str]]:
        success = 0
        errors: List[str] = []
        for idx, row in enumerate(rows, start=1):
            try:
                email = row.get("email")
                code = row.get("student_code") or row.get("code")
                first = row.get("first_name") or row.get("prenom")
                last = row.get("last_name") or row.get("nom")
                class_name = row.get("class") or row.get("class_name") or row.get("classe")
                if not all([email, code, first, last, class_name]):
                    raise ValueError("Missing required student fields")

                user = ImportService._ensure_user(
                    db, email=email, first_name=first, last_name=last, role="student"
                )

                student = db.query(Student).filter(Student.student_code == code).first()
                if not student:
                    student = Student(student_code=code, user_id=user.id)
                    db.add(student)

                student.first_name = first
                student.last_name = last
                student.email = email
                student.class_name = class_name
                student.group_name = row.get("group") or row.get("group_name") or student.group_name
                student.academic_status = row.get("academic_status") or student.academic_status
                student.enrollment_date = (
                    ImportService._parse_date(row.get("enrollment_date")) or student.enrollment_date
                )
                student.expected_graduation = (
                    ImportService._parse_date(row.get("expected_graduation"))
                    or student.expected_graduation
                )
                success += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"row {idx}: {exc}")
        db.commit()
        return success, errors

    @staticmethod
    def import_trainers(db: Session, rows: Iterable[Dict[str, str]]) -> Tuple[int, List[str]]:
        success = 0
        errors: List[str] = []
        for idx, row in enumerate(rows, start=1):
            try:
                email = row.get("email")
                first = row.get("first_name") or row.get("prenom")
                last = row.get("last_name") or row.get("nom")
                specialization = row.get("specialization") or row.get("specialite")
                if not all([email, first, last]):
                    raise ValueError("Missing required trainer fields")

                user = ImportService._ensure_user(
                    db, email=email, first_name=first, last_name=last, role="trainer"
                )

                trainer = db.query(Trainer).filter(Trainer.email == email).first()
                if not trainer:
                    trainer = Trainer(email=email, user_id=user.id)
                    db.add(trainer)
                trainer.first_name = first
                trainer.last_name = last
                trainer.specialization = specialization
                trainer.status = row.get("status") or trainer.status
                success += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"row {idx}: {exc}")
        db.commit()
        return success, errors

    @staticmethod
    def import_sessions(db: Session, rows: Iterable[Dict[str, str]]) -> Tuple[int, List[str]]:
        success = 0
        errors: List[str] = []
        for idx, row in enumerate(rows, start=1):
            try:
                module_id = ImportService._as_int(row.get("module_id"))
                trainer_id = ImportService._as_int(row.get("trainer_id"))
                classroom_id = ImportService._as_int(row.get("classroom_id"))
                session_date = ImportService._parse_date(row.get("session_date"))
                start_time = ImportService._parse_time(row.get("start_time"))
                end_time = ImportService._parse_time(row.get("end_time"))
                if not all(
                    [module_id, trainer_id, classroom_id, session_date, start_time, end_time]
                ):
                    raise ValueError("Missing required session fields")

                session = SessionModel(
                    module_id=module_id,
                    trainer_id=trainer_id,
                    classroom_id=classroom_id,
                    session_date=session_date,
                    start_time=start_time,
                    end_time=end_time,
                    status=row.get("status") or "scheduled",
                    session_type=row.get("session_type") or "theory",
                )
                db.add(session)
                success += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"row {idx}: {exc}")
        db.commit()
        return success, errors

    # ---------- Helpers ----------
    @staticmethod
    def _parse_date(raw: str | None) -> date | None:
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_time(raw: str | None) -> time | None:
        if not raw:
            return None
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(raw, fmt).time()
            except ValueError:
                continue
        return None

    @staticmethod
    def _as_int(raw: str | None) -> int | None:
        try:
            return int(raw) if raw is not None and raw != "" else None
        except ValueError:
            return None

    @staticmethod
    def templates() -> Dict[str, str]:
        """Return CSV template content for all entities."""
        return {
            "students": "student_code,first_name,last_name,email,class_name,group_name,academic_status,enrollment_date,expected_graduation\nSTU-001,Jane,Doe,jane@example.com,LICENCE 1,A,active,2024-09-01,2027-06-30\n",
            "trainers": "first_name,last_name,email,specialization,status\nAlex,Trainer,alex@example.com,Math,active\n",
            "sessions": "module_id,trainer_id,classroom_id,session_date,start_time,end_time,session_type,status\n12,3,7,2025-01-15,09:00,11:00,theory,scheduled\n",
        }
