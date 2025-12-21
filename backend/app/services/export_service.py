"""
Export Service - Generate PDF and Excel reports
"""

import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.attendance import AttendanceRecord
from app.models.session import Session as CourseSession
from app.models.smart_attendance import AttendanceAlert
from app.models.student import Student


class ExportService:
    """Service for exporting attendance and analytics data."""
    
    @staticmethod
    def export_attendance_pdf(
        db: Session,
        session_id: int,
        include_stats: bool = True,
    ) -> bytes:
        """
        Export attendance report as PDF.
        
        Returns PDF file bytes.
        """
        
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        
        # Get session info
        session = db.query(CourseSession).filter(CourseSession.id == session_id).first()
        if not session:
            raise ValueError("Session not found")
        
        # Title
        title = Paragraph("Rapport de Présence", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2 * inch))
        
        # Session info
        session_date = getattr(session, "session_date", None)
        session_topic = getattr(session, "topic", None) or getattr(session, "title", None)

        session_info = [
            ["Session:", session_topic or "N/A"],
            ["Date:", session_date.strftime("%d/%m/%Y") if session_date else "N/A"],
            ["Heure:", f"{session.start_time} - {session.end_time}" if session.start_time else "N/A"],
            ["Classe:", session.class_name or "N/A"],
        ]
        
        info_table = Table(session_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Get attendance records
        attendance_records = db.query(AttendanceRecord, Student).join(
            Student, AttendanceRecord.student_id == Student.id
        ).filter(
            AttendanceRecord.session_id == session_id
        ).all()
        
        # Statistics
        if include_stats:
            total = len(attendance_records)
            present = sum(1 for r, _ in attendance_records if r.status == "present")
            absent = sum(1 for r, _ in attendance_records if r.status == "absent")
            late = sum(1 for r, _ in attendance_records if r.status == "late")
            
            stats_title = Paragraph("<b>Statistiques</b>", styles['Heading2'])
            story.append(stats_title)
            story.append(Spacer(1, 0.1 * inch))
            
            stats_data = [
                ["Total étudiants", str(total)],
                ["Présents", f"{present} ({present/total*100:.1f}%)" if total > 0 else "0"],
                ["Absents", f"{absent} ({absent/total*100:.1f}%)" if total > 0 else "0"],
                ["Retards", f"{late} ({late/total*100:.1f}%)" if total > 0 else "0"],
            ]
            
            stats_table = Table(stats_data, colWidths=[2*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 0.3 * inch))
        
        # Attendance table
        attendance_title = Paragraph("<b>Liste de Présence</b>", styles['Heading2'])
        story.append(attendance_title)
        story.append(Spacer(1, 0.1 * inch))
        
        attendance_data = [["#", "Nom", "Prénom", "N° Étudiant", "Statut"]]
        
        for idx, (record, student) in enumerate(attendance_records, 1):
            status_text = {
                "present": "Présent",
                "absent": "Absent",
                "late": "Retard",
            }.get(record.status, record.status)

            student_number = getattr(student, "student_number", None) or getattr(student, "student_code", None)
            
            attendance_data.append([
                str(idx),
                student.last_name or "",
                student.first_name or "",
                student_number or "",
                status_text,
            ])
        
        attendance_table = Table(attendance_data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1*inch])
        attendance_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(attendance_table)
        
        # Footer
        story.append(Spacer(1, 0.5 * inch))
        footer = Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} par SmartPresence",
            styles['Normal']
        )
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def export_attendance_excel(
        db: Session,
        session_id: Optional[int] = None,
        student_id: Optional[int] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> bytes:
        """
        Export attendance data as Excel file.
        
        More flexible than PDF - can export multiple sessions or by student.
        Returns Excel file bytes.
        """
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Présence"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Date", "Session", "Étudiant", "N° Étudiant", "Statut", "Heure", "Méthode"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Build query
        query = db.query(AttendanceRecord, Student, CourseSession).join(
            Student, AttendanceRecord.student_id == Student.id
        ).join(
            CourseSession, AttendanceRecord.session_id == CourseSession.id
        )
        
        if session_id:
            query = query.filter(AttendanceRecord.session_id == session_id)
        if student_id:
            query = query.filter(AttendanceRecord.student_id == student_id)
        if start_date:
            query = query.filter(CourseSession.session_date >= start_date)
        if end_date:
            query = query.filter(CourseSession.session_date <= end_date)
        
        records = query.all()
        
        # Add data
        for record, student, session in records:
            session_date = getattr(session, "session_date", None)
            session_topic = getattr(session, "topic", None) or getattr(session, "title", None)
            student_number = getattr(student, "student_number", None) or getattr(student, "student_code", None)
            row = [
                session_date.strftime("%d/%m/%Y") if session_date else "",
                session_topic or "",
                f"{student.last_name} {student.first_name}",
                student_number or "",
                record.status or "",
                record.marked_at.strftime("%H:%M") if record.marked_at else "",
                record.marked_via or "",
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add border to all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    @staticmethod
    def export_alerts_excel(
        db: Session,
        severity: Optional[str] = None,
        acknowledged: Optional[bool] = None,
    ) -> bytes:
        """Export alerts as Excel."""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertes"
        
        # Headers
        headers = ["ID", "Étudiant", "Type", "Gravité", "Message", "Date", "Acquitté", "Action"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Query alerts
        query = db.query(AttendanceAlert, Student).join(
            Student, AttendanceAlert.student_id == Student.id
        )
        
        if severity:
            query = query.filter(AttendanceAlert.severity == severity)
        if acknowledged is not None:
            query = query.filter(AttendanceAlert.is_acknowledged == acknowledged)
        
        alerts = query.all()
        
        # Add data
        for alert, student in alerts:
            row = [
                alert.id,
                f"{student.last_name} {student.first_name}",
                alert.alert_type or "",
                alert.severity or "",
                alert.message or "",
                alert.created_at.strftime("%d/%m/%Y %H:%M") if alert.created_at else "",
                "Oui" if alert.is_acknowledged else "Non",
                alert.action_taken or "",
            ]
            ws.append(row)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
